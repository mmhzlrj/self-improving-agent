#!/usr/bin/env python3
"""
Excel 处理工具
用法: python3 excel.py --input data.xlsx --output result.xlsx
"""

import openpyxl
import argparse
import sys

def read_excel(file_path, sheet=0):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[sheet]
    
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    return data

def write_excel(file_path, data, headers=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if headers:
        ws.append(headers)
    
    for row in data:
        ws.append(row)
    
    wb.save(file_path)
    print(f"✓ 已保存到 {file_path}")

def add_column(file_path, col_name, values):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 添加表头
    max_col = ws.max_column + 1
    ws.cell(1, max_col, col_name)
    
    # 添加数据
    for i, val in enumerate(values, 2):
        ws.cell(i, max_col, val)
    
    wb.save(file_path)
    print(f"✓ 已添加列: {col_name}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Excel 处理工具')
    parser.add_argument('--input', '-i', required=True, help='输入文件')
    parser.add_argument('--output', '-o', help='输出文件')
    parser.add_argument('--add-col', help='添加列名')
    parser.add_argument('--values', nargs='+', help='列值')
    
    args = parser.parse_args()
    
    if args.add_col and args.values:
        add_column(args.input, args.add_col, args.values)
    elif args.output:
        data = read_excel(args.input)
        write_excel(args.output, data)
    else:
        data = read_excel(args.input)
        for row in data[:10]:
            print(row)
